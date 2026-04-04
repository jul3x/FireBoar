import flet as ft
import json
import re
from datetime import datetime
import asyncio
import httpx
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Side, Border, PatternFill, Font
from io import BytesIO
from fireboar.training import Training, Exercise, Session
from fireboar.storage import save_trainings, save_sessions, load_trainings, load_sessions
from fireboar.utils import show_dialog, normalize_string


async def import_json(page, files):
    if not files:
        return

    file = files[0]
    if not file.bytes:
        await show_dialog(
            page,
            "Coś nie pykło",
            "Wrzuciłeś mi pusty plik?",
            "Oczy widzą, usta milczą",
        )
        return

    try:
        file_bytes = file.bytes.decode('utf-8')
        data = json.loads(file_bytes)
        trainings = [Training.from_json(t) for t in data["trainings"]]
        sessions = [Session.from_json(s) for s in data["sessions"]]
        await save_trainings(trainings)
        await save_sessions(sessions)
    except Exception as e:
        print(e)
        await show_dialog(
            page,
            "Taki cwany?",
            "Wrzuć mi coś normalnego, to było zepsute.",
            "Ok",
        )
        return

    await show_dialog(
        page,
        "Dane zaimportowane",
        "Poprzednie dane zostały wyczyszczone!",
        "Ok",
    )


async def export_json(json_file_picker):
    trainings = [t.to_json() for t in await load_trainings()]
    sessions = [s.to_json() for s in await load_sessions()]
    curr_date = datetime.now().strftime("%Y_%m_%d")
    await json_file_picker.save_file(
        dialog_title="Zapisz treningi",
        file_name=f"fireboar_trainings_{curr_date}.json",
        src_bytes=json.dumps({"trainings": trainings, "sessions": sessions}).encode("utf-8"),
    )


async def export_kate(file_picker):
    trainings = await load_trainings()
    sessions = await load_sessions()
    curr_date = datetime.now().strftime("%Y_%m_%d")
    sheet_data = await get_training_sheet(trainings, sessions)
    await file_picker.save_file(
        dialog_title="Zapisz arkusz treningowy",
        file_name=f"fireboar_trainings_{curr_date}.xlsx",
        src_bytes=sheet_data,
    )


async def import_kate_entry(page, home_function):
    page.controls.clear()
    page.add(ft.Text("Wpisz URL do arkusza Google (tfu)", size=24, weight="bold", width=4000, text_align="center"))
    page.add(ft.Text("Upewnij się, że masz dostęp do odczytu za pomocą linka.", size=16, width=4000, text_align="center"))
    page.add(
        url_field := ft.TextField(
            label="URL",
            expand=True,
            border_color="#555555",
            color="#ffffff",
            bgcolor="#111111",
        ),
    )

    exit_event = asyncio.Event()
    async def import_xlsx(e):
        button.content = "wczytuję..."
        button.disabled = True
        page.update()
        url = url_field.value
        try:
            url = url.strip().split("/")
            url = "/".join(url[:-1])
            url += '/export?format=xlsx'
            async with httpx.AsyncClient(follow_redirects=True) as client:
                r = await client.get(url)
                file_content = r.content

            await import_kate(page, file_content, home_function)
        except Exception as ex:
            await show_dialog(
                page,
                "Coś nie pykło",
                "Czy to na pewno URL skopiowany prosto od wujka Google?",
                "Ogar",
            )
            print(ex)
            pass

        button.content = "Wczytaj arkusz"
        button.disabled = False
        page.update()
        await exit_import(e)

    async def exit_import(e):
        exit_event.set()

    page.add(button := ft.Button("Wczytaj arkusz", width=4000, height=50, on_click=import_xlsx))
    page.add(ft.TextButton("Wróć", width=4000, height=50, on_click=exit_import))
    page.update()
    await exit_event.wait()



async def import_kate(page, file_content, home_function):
    if not file_content:
        await show_dialog(
            page,
            "Coś nie pykło",
            "Wrzuciłeś mi pusty plik?",
            "Oczy widzą, usta milczą",
        )
        return

    try:
        file_memory = BytesIO(file_content)
        wb = load_workbook(file_memory)
    except Exception as e:
        print(e)
        await show_dialog(
            page,
            "Taki cwany?",
            "Wrzuć mi coś normalnego, to było zepsute.",
            "Ok",
        )
        return

    exit_event = asyncio.Event()
    sheet_picker = ft.RadioGroup(
        expand=True,
        content=ft.Column([
            ft.Radio(value=sheet_name, label=sheet_name)
            for sheet_name in wb.sheetnames
        ]),
        on_change=lambda e: None,
    )
    page.controls.clear()
    page.add(ft.Text("Plik załadowany.", size=24, weight="bold", width=4000, text_align="center"))
    page.add(ft.Text("Który trening chcesz zaimportować?", size=20, width=4000, text_align="center"))
    page.add(sheet_picker)

    async def continue_import():
        training = await process_kate_sheet(page, wb, sheet_picker.value)
        exit_event.set()

    async def exit():
        exit_event.set()

    page.add(
        ft.Row([
            ft.Button("Kontynuuj", on_click=continue_import),
            ft.TextButton("⬅ Wróć", on_click=exit)
        ]),
    )

    await exit_event.wait()


async def process_kate_sheet(page, wb, sheet_name: str | None) -> Training:
    if sheet_name not in wb.sheetnames:
        await show_dialog(
            page,
            "Zły wybór",
            "Spróbuj ponownie byczq!",
            "Ok",
        )
        return

    sheet = wb[sheet_name]
    rows = list(sheet.iter_rows())
    training = Training(name=sheet_name)
    for idx, row in enumerate(rows):
        if not row[0] or not row[0].value:
            continue

        if "TRENING" == str(row[0].value).strip().upper():
            try:
                sets = str(rows[idx + 1][1].value or 1)
                sets = int(float(normalize_string(sets.split("L")[0].split("P")[0] or "")))

                reps_to_parse = rows[idx + 1][2].value
                # Weird but valid
                if isinstance(reps_to_parse, datetime):
                    suggested_reps = f"{reps_to_parse.day} - {reps_to_parse.month}"
                else:
                    suggested_reps = str(reps_to_parse or "")
                suggested_weight = str(rows[idx + 1][3].value or "")

                header_idx = idx - 1
                while header_idx >= 0 and idx - header_idx < 7 and not rows[header_idx][0].value:
                    header_idx -= 1

                superset = str(rows[header_idx][0].value or "")
                superset = re.sub(r"[0-9]", "", superset)
                rest = get_rest(rows[header_idx][0])
                names = rows[header_idx][1].value.split('\n')
                parsed_names = []
                if "core" in superset.lower().strip():
                    for name in names:
                        for n in name.split(','):
                            n = str(n.split('http')[0])
                            parsed_names.append(n.strip().strip(':'))
                else:
                    for name in names:
                        name = str(name.split('http')[0])
                        parsed_names.append(name.strip().strip(':'))

                for name in parsed_names:
                    exercise = Exercise(
                        name=name,
                        sets=sets,
                        suggested_weight=suggested_weight,
                        suggested_reps=suggested_reps,
                        rest_seconds=rest,
                        superset_id=superset,
                    )
                    training.exercises.append(exercise)
            except:
                pass

    trainings = await load_trainings()
    trainings = [training] + trainings
    await save_trainings(trainings)
    await show_dialog(
        page,
        "Trening dodany",
        "Zobacz proszę, czy wszystko zostało dodane odpowiednio. Zwróć uwagę szczególnie na ćwiczenia CORE i ćwiczenia, które powinny być interwałowymi (podciąganie klastrowe / chwyto). Sprawdź czy resty są ustawione odpowiednio.",
        "Ładuj",
    )


async def get_training_sheet(trainings: list[Training], sessions: list[Session]) -> bytes:
    buffer = BytesIO()
    wb = Workbook()
    wb.remove(wb.active)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for t in trainings:
        wb.create_sheet(t.name)
        sheet = wb[t.name]

        curr_row = 1
        starting_row = 0
        supersets = {}

        sheet.column_dimensions['A'].width = 12
        sheet.column_dimensions['D'].width = 10
        sheet.column_dimensions['E'].width = 25
        sheet.column_dimensions['F'].width = 10
        sheet.column_dimensions['G'].width = 25
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['I'].width = 25
        sheet.column_dimensions['J'].width = 10
        sheet.column_dimensions['K'].width = 25
        s_for_training = t.get_sessions(sessions)
        for ex in sorted(t.exercises, key=lambda e: e.superset_id):
            bgfill = get_bg_for_exercise(ex)

            supersets.setdefault(ex.superset_id, 0)
            supersets[ex.superset_id] += 1

            # Header
            sheet[f"A{curr_row}"].fill = bgfill
            if ex.superset_id:
                sheet[f"A{curr_row}"] = f"{ex.superset_id}{supersets[ex.superset_id]}"
            sheet[f"B{curr_row}"] = ex.name
            sheet[f"B{curr_row}"].fill = bgfill

            # Second headers
            curr_row += 1
            sheet[f"A{curr_row}"] = "TRENING"
            sheet[f"B{curr_row}"] = "SERIE"
            sheet[f"C{curr_row}"] = "POWT."

            curr_column = 'C'
            for s in range(ex.sets):
                curr_column = chr(ord(curr_column) + 1)
                sheet[f"{curr_column}{curr_row}"] = f"SERIA {s + 1}"
                curr_column = chr(ord(curr_column) + 1)
                sheet[f"{curr_column}{curr_row}"] = f"uwagi"

            sheet.merge_cells(f"B{curr_row - 1}:{curr_column}{curr_row - 1}")
            max_column = curr_column

            # Data
            curr_row += 1
            for s in s_for_training:
                sheet[f"A{curr_row}"] = s.get_date()
                sheet[f"B{curr_row}"] = str(ex.sets)
                sheet[f"C{curr_row}"] = str(ex.suggested_reps)
                curr_column = 'C'
                for session_set in s.sets:
                    if session_set.get_id() != ex.id:
                        continue

                    curr_column = chr(ord(curr_column) + 1)
                    sheet[f"{curr_column}{curr_row}"] = f"{session_set.weight} x {session_set.reps}"
                    curr_column = chr(ord(curr_column) + 1)
                    sheet[f"{curr_column}{curr_row}"] = session_set.notes
                curr_row += 1

            for row in sheet.iter_rows(min_row=starting_row + 1, max_row=curr_row - 1, min_col=1, max_col=ord(max_column) - ord('A') + 1):
                for cell in row:
                    cell.border = thin_border

            starting_row = curr_row
            curr_row += 1

        font = Font(name="Roboto Mono", size=12, bold=False, color="000000")
        for row in sheet.iter_rows():
            for cell in row:
                cell.font = font

    wb.save(buffer)
    return buffer.getvalue()


LONG_REST_COLOR = (74, 134, 232)
MEDIUM_REST_COLOR = (201, 218, 248)
SHORT_REST_COLOR = (217, 234, 211)


def get_bg_for_exercise(ex: Exercise) -> PatternFill:
    long_rest = '{:02X}{:02X}{:02X}'.format(*LONG_REST_COLOR)
    medium_rest = '{:02X}{:02X}{:02X}'.format(*MEDIUM_REST_COLOR)
    short_rest = '{:02X}{:02X}{:02X}'.format(*SHORT_REST_COLOR)

    if ex.rest_seconds < 40:
        color = short_rest
    if ex.rest_seconds >= 40:
        color = medium_rest
    if ex.rest_seconds > 80:
        color = long_rest
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def excel_rgb_to_tuple(argb):
    if argb is None:
        return None
    argb = argb[-6:]  # remove alpha
    return tuple(int(argb[i:i+2], 16) for i in (0, 2, 4))


def get_rest(cell) -> int:
    fill = cell.fill
    color = fill.fgColor.rgb

    rgb = excel_rgb_to_tuple(color)

    def color_distance_sq(c1, c2):
        return sum((a - b) ** 2 for a, b in zip(c1, c2))

    targets = {
        180: LONG_REST_COLOR,
        60: MEDIUM_REST_COLOR,
        20: SHORT_REST_COLOR,
    }

    closest = min(
        targets.items(),
        key=lambda item: color_distance_sq(rgb, item[1])
    )

    return closest[0]
